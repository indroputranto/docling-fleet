#!/usr/bin/env python3
"""
Packing-list spreadsheet parser.

Accepts .xlsx (openpyxl) and .xls (xlrd==1.2.0) files and returns a normalized
list of cargo-item dicts ready for insertion into the `cargo_items` table.

Why a custom parser instead of pandas read_excel?
  - The two real-world packing lists we support both prepend several rows of
    free-text metadata (company name, shipment ref, vessel, contract number…)
    above the actual data table.  We need to *find* the header row, not assume
    it's row 1.
  - Column names vary across templates (English vs. Italian, abbreviated vs.
    full) so we score candidate header rows against a vocabulary and pick the
    best match — this beats hardcoded column indexes.
  - We need to gracefully skip totals rows ("TOT.", "TOTALE", "TOTAL") and
    fully-blank rows that some files leave between the table and trailing
    notes.

All output dimensions are normalized to METERS, weights to KILOGRAMS,
volume to CUBIC METERS.  When the source file uses millimeters or tonnes
(detected via header text or magnitude heuristics) we convert.

Public API:
    parse_packing_list(path_or_bytes, filename=None) -> ParseResult

Returns ParseResult dataclass with:
    items:   list[dict]   normalized cargo items
    header:  list[str]    detected header row (raw text, after normalization)
    skipped: list[dict]   rows we couldn't parse (with reason)
    file_type: str        "xlsx" | "xls"
    sheet_name: str       sheet that was parsed
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
from dataclasses import dataclass, field
from typing import Any, Iterable, Optional, Union

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Column vocabulary
# ─────────────────────────────────────────────────────────────────────────────

# Each entry maps a normalized output field → list of regex patterns matched
# against (lowercased, punctuation-stripped) header cells.  First match wins.
# Order patterns from most-specific to most-general so e.g. "GROSS WEIGHT"
# beats "WEIGHT".
_COLUMN_PATTERNS: dict[str, list[str]] = {
    "item_id": [
        r"\bpl\s*n[°o]?\b",            # "PL N°" (file 1)
        r"\bnr\.?\s*plist\b",           # "NR. PLIST" (file 2)
        r"\bpacking\s*list\s*(no|number|id)\b",
        r"\bpl\s*id\b",
        r"\bitem\s*(id|no|number|code)\b",
        r"\bpart\s*(no|number)\b",
        r"\bposition\s*(no|number)\b",
        r"\bcase\s*(no|number)\b",
    ],
    "description": [
        r"\bbox\s*description\b",       # file 1: "Box Description"
        r"\bpei\s*description\b",       # file 1: "PEI Description"
        r"\bdescription\b",
        r"\bdescriz\b",                  # Italian: "DESCRIZ."
        r"\bdescrizione\b",
        r"\bcommodity\b",
        r"\bgoods\b",
        r"\bcomm\.?\b",                  # file 2: "COMM."
    ],
    "packing_type": [
        r"\bpacking\s*type\b",
        r"\bpackage\s*type\b",
        r"\bpkg\s*type\b",
        r"\btype\s*of\s*pack\w*\b",
    ],
    "length_m": [
        r"\bdim\w*\s*[\[\(]?m[m\)\]]?\s*l\b",      # "Dimensions [m] L"
        r"\blength\b",
        r"\blunghezza\b",
        r"^l$",
        r"^l\s*[\[\(]m[m]?[\]\)]\s*$",
    ],
    "width_m": [
        r"\bdim\w*\s*[\[\(]?m[m\)\]]?\s*w\b",
        r"\bwidth\b",
        r"\blarghezza\b",
        r"^w$",
        r"^w\s*[\[\(]m[m]?[\]\)]\s*$",
    ],
    "height_m": [
        r"\bdim\w*\s*[\[\(]?m[m\)\]]?\s*h\b",
        r"\bheight\b",
        r"\baltezza\b",
        r"^h$",
        r"^h\s*[\[\(]m[m]?[\]\)]\s*$",
    ],
    "volume_m3": [
        r"\bvolume\s*m\s*3\b",
        r"\bvolume\b",
        r"\bcbm\b",                     # cubic meters
        r"\bm\s*3\b",
        r"\bvol\.?\s*[\[\(]?m\s*3[\]\)]?\b",
    ],
    "net_weight_kg": [
        r"\bnet\s*weight\b",
        r"\bn\.?\s*w\.?\b",
        r"\bpeso\s*netto\b",
    ],
    "gross_weight_kg": [
        r"\bgross\s*weight\b",
        r"\bg\.?\s*w\.?\b",
        r"\bpeso\s*lordo\b",
    ],
    "imo_flag": [
        r"\bimo(\s*[\(\[]?\s*y/?n\s*[\)\]]?)?\b",
        r"\bhazmat\b",
        r"\bimdg\b",
        r"\bdangerous\b",
    ],
}


# Header-row scoring: minimum number of recognized columns required for a row
# to qualify as the header.  We need ID + at least 2 of (L, W, H, weight) to
# call it confidently.
_MIN_RECOGNIZED_COLUMNS = 4


# Tokens that mark a row as a "totals" / footer row to be skipped.
_TOTALS_TOKENS = {"tot", "total", "totale", "totals", "totali", "subtotal", "sum"}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_header_cell(cell: Any) -> str:
    """Lowercase, collapse whitespace, strip surrounding punctuation."""
    if cell is None:
        return ""
    s = str(cell).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _match_column(header_cell_norm: str) -> Optional[str]:
    """Return the canonical field name for a header cell, or None."""
    if not header_cell_norm:
        return None
    for field_name, patterns in _COLUMN_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, header_cell_norm, flags=re.I):
                return field_name
    return None


def _score_header_row(row: list[Any]) -> tuple[int, dict[int, str]]:
    """Score a candidate header row. Returns (score, {col_index: field_name})."""
    mapping: dict[int, str] = {}
    seen_fields: set[str] = set()
    for idx, cell in enumerate(row):
        norm = _normalize_header_cell(cell)
        field_name = _match_column(norm)
        if field_name and field_name not in seen_fields:
            mapping[idx] = field_name
            seen_fields.add(field_name)
    # Bonus weight if we have at least an ID + 2 of (L, W, H)
    score = len(mapping)
    if "item_id" in seen_fields and \
       sum(f in seen_fields for f in ("length_m", "width_m", "height_m")) >= 2:
        score += 3
    if "gross_weight_kg" in seen_fields:
        score += 1
    return score, mapping


def _to_float(val: Any) -> Optional[float]:
    """Coerce a cell value to float, handling commas, units, blanks."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    # Strip unit suffixes (kg, m, mm, m3, cbm, etc.)
    s = re.sub(r"[a-zA-Z]+\s*$", "", s).strip()
    # Italian-style decimal comma
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    elif "," in s and "." in s:
        # Treat comma as thousands separator
        s = s.replace(",", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _looks_like_totals_row(row: list[Any]) -> bool:
    """Return True if any of the first 3 cells contains a totals keyword."""
    for cell in row[:3]:
        if cell is None:
            continue
        s = str(cell).strip().lower()
        # split on punctuation/spaces and check tokens
        tokens = set(re.split(r"[\s\.\,:;\-]+", s))
        if tokens & _TOTALS_TOKENS:
            return True
    return False


def _is_imo_truthy(val: Any) -> bool:
    """Interpret an IMO-flag cell as boolean. 'Y', 'YES', 'TRUE', '1' = True."""
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("Y", "YES", "TRUE", "1", "X", "IMO", "IMDG", "DG", "DANGEROUS")


def _maybe_convert_dimension(val: float, header_text: str) -> float:
    """
    If the header indicates millimeters, convert mm→m. Otherwise return as-is.
    Also catch obviously-mm values (>50 m would be a 50m+ box; very unlikely)
    as a magnitude fallback.
    """
    if val is None:
        return 0.0
    h = (header_text or "").lower()
    if "mm" in h or "millim" in h:
        return val / 1000.0
    if "cm" in h or "centim" in h:
        return val / 100.0
    # Magnitude heuristic: dims > 30m almost certainly mm
    if val > 30:
        return val / 1000.0
    return val


def _maybe_convert_weight(val: float, header_text: str) -> float:
    """Convert tonnes→kg if header indicates t / tonne."""
    if val is None:
        return 0.0
    h = (header_text or "").lower()
    if re.search(r"\b(t|tonne|metric\s*ton)\b", h) and "[kg]" not in h and "kg" not in h:
        return val * 1000.0
    return val


def _row_to_list(row, n_cols: int) -> list[Any]:
    """Convert sheet row (openpyxl tuple or xlrd list) to fixed-width list."""
    cells = list(row[:n_cols]) if not isinstance(row, list) else row[:n_cols]
    while len(cells) < n_cols:
        cells.append(None)
    return cells


# ─────────────────────────────────────────────────────────────────────────────
# ParseResult
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ParseResult:
    items: list[dict]               = field(default_factory=list)
    header: list[str]               = field(default_factory=list)
    column_mapping: dict[int, str]  = field(default_factory=dict)
    skipped: list[dict]             = field(default_factory=list)
    file_type: str                  = ""
    sheet_name: str                 = ""
    header_row_idx: int             = 0

    def to_dict(self) -> dict:
        return {
            "items":          self.items,
            "header":         self.header,
            "column_mapping": {str(k): v for k, v in self.column_mapping.items()},
            "skipped":        self.skipped,
            "file_type":      self.file_type,
            "sheet_name":     self.sheet_name,
            "header_row_idx": self.header_row_idx,
        }


# ─────────────────────────────────────────────────────────────────────────────
# Sheet readers
# ─────────────────────────────────────────────────────────────────────────────

def _read_xlsx(stream: io.BytesIO) -> tuple[str, list[list[Any]]]:
    """Return (sheet_name, rows) using openpyxl. Picks the largest sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(stream, data_only=True, read_only=True)
    best_sheet = None
    best_rowcount = -1
    for sn in wb.sheetnames:
        ws = wb[sn]
        # ws.max_row can lie in read-only mode; iterate to count
        n = ws.max_row or 0
        if n > best_rowcount:
            best_rowcount = n
            best_sheet = sn

    if best_sheet is None:
        return "", []

    ws = wb[best_sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return best_sheet, rows


def _read_xls(stream: io.BytesIO) -> tuple[str, list[list[Any]]]:
    """Return (sheet_name, rows) using xlrd 1.2.0."""
    try:
        import xlrd  # xlrd==1.2.0 — newer versions dropped .xls support
    except ImportError as exc:
        raise RuntimeError(
            "Reading legacy .xls files requires the xlrd package. "
            "Run:  pip install 'xlrd==1.2.0'  (newer versions of xlrd "
            "dropped binary .xls support; pin to 1.2.0). "
            "Alternatively, re-save the file as .xlsx in Excel and re-upload."
        ) from exc

    data = stream.read() if hasattr(stream, "read") else stream
    book = xlrd.open_workbook(file_contents=data)
    best_sheet = None
    best_rowcount = -1
    for sheet in book.sheets():
        if sheet.nrows > best_rowcount:
            best_rowcount = sheet.nrows
            best_sheet = sheet
    if best_sheet is None:
        return "", []

    rows: list[list[Any]] = []
    for r in range(best_sheet.nrows):
        rows.append([best_sheet.cell_value(r, c) for c in range(best_sheet.ncols)])
    return best_sheet.name, rows


# ─────────────────────────────────────────────────────────────────────────────
# Header detection
# ─────────────────────────────────────────────────────────────────────────────

def _find_header_row(rows: list[list[Any]]) -> tuple[int, dict[int, str]]:
    """
    Scan the first 30 rows for the best-scoring header row.
    Returns (row_index, column_mapping). Raises ValueError if none found.
    """
    best_idx = -1
    best_score = 0
    best_mapping: dict[int, str] = {}

    scan_limit = min(30, len(rows))
    for i in range(scan_limit):
        row = rows[i]
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        score, mapping = _score_header_row(row)
        if score > best_score:
            best_score = score
            best_idx = i
            best_mapping = mapping

    if best_idx < 0 or best_score < _MIN_RECOGNIZED_COLUMNS:
        raise ValueError(
            f"Could not locate a packing-list header row in the first "
            f"{scan_limit} rows (best score={best_score}, "
            f"need ≥ {_MIN_RECOGNIZED_COLUMNS} recognized columns)."
        )
    return best_idx, best_mapping


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def parse_packing_list(
    path_or_bytes: Union[str, bytes, io.BytesIO],
    filename: Optional[str] = None,
) -> ParseResult:
    """
    Parse a packing list and return normalized cargo items.

    Parameters
    ----------
    path_or_bytes : str | bytes | BytesIO
        File on disk, raw bytes, or a buffered binary stream.
    filename : str, optional
        Used solely to detect file format when bytes/stream is passed.
        Required when *path_or_bytes* is bytes/stream.
    """
    # ── Resolve file type and stream ────────────────────────────────────────
    if isinstance(path_or_bytes, str):
        fname = filename or os.path.basename(path_or_bytes)
        with open(path_or_bytes, "rb") as f:
            buf = io.BytesIO(f.read())
    elif isinstance(path_or_bytes, (bytes, bytearray)):
        if not filename:
            raise ValueError("filename is required when passing raw bytes")
        fname = filename
        buf = io.BytesIO(path_or_bytes)
    else:  # assume BytesIO-like
        if not filename:
            raise ValueError("filename is required when passing a stream")
        fname = filename
        buf = path_or_bytes

    ext = (os.path.splitext(fname)[1] or "").lower().lstrip(".")
    if ext not in ("xlsx", "xlsm", "xls"):
        raise ValueError(
            f"Unsupported packing-list file type: '.{ext}'. "
            "Expected one of: xlsx, xlsm, xls."
        )

    file_type = "xls" if ext == "xls" else "xlsx"

    # ── Read sheet ──────────────────────────────────────────────────────────
    if file_type == "xls":
        sheet_name, rows = _read_xls(buf)
    else:
        sheet_name, rows = _read_xlsx(buf)

    if not rows:
        raise ValueError(f"Spreadsheet '{fname}' is empty.")

    # ── Find header ─────────────────────────────────────────────────────────
    header_idx, mapping = _find_header_row(rows)
    header_row = rows[header_idx]
    n_cols = max(len(header_row), max(mapping.keys()) + 1 if mapping else 1)

    # Capture raw header text per recognized field for unit detection later
    field_to_header_text: dict[str, str] = {
        f: _normalize_header_cell(header_row[i]) for i, f in mapping.items()
    }

    # ── Iterate data rows ───────────────────────────────────────────────────
    items: list[dict] = []
    skipped: list[dict] = []

    for r in range(header_idx + 1, len(rows)):
        raw_row = _row_to_list(rows[r], n_cols)

        # Empty row → stop reading (most files end with blank tail)
        if all(c is None or str(c).strip() == "" for c in raw_row):
            continue

        if _looks_like_totals_row(raw_row):
            skipped.append({"row": r, "reason": "totals row", "raw": [str(c) for c in raw_row[:5]]})
            continue

        # Build the normalized item
        item: dict[str, Any] = {
            "position":              len(items),
            "item_id":               None,
            "description":           None,
            "packing_type":          None,
            "length_m":              0.0,
            "width_m":               0.0,
            "height_m":              0.0,
            "volume_m3":             0.0,
            "net_weight_kg":         None,
            "gross_weight_kg":       0.0,
            "imo_flag":              False,
            "can_stack":             True,
            "can_rotate_horizontal": True,
        }
        raw_row_dump: dict[str, Any] = {}

        for col_idx, field_name in mapping.items():
            cell = raw_row[col_idx] if col_idx < len(raw_row) else None
            raw_row_dump[field_name] = (
                cell.isoformat() if hasattr(cell, "isoformat") else cell
            )
            header_text = field_to_header_text.get(field_name, "")

            if field_name == "item_id":
                item["item_id"] = str(cell).strip() if cell not in (None, "") else None
            elif field_name == "description":
                if cell not in (None, ""):
                    txt = str(cell).strip()
                    item["description"] = (item["description"] + " | " + txt) \
                        if item["description"] else txt
            elif field_name == "packing_type":
                if cell not in (None, ""):
                    item["packing_type"] = str(cell).strip()
            elif field_name in ("length_m", "width_m", "height_m"):
                v = _to_float(cell)
                if v is not None:
                    item[field_name] = _maybe_convert_dimension(v, header_text)
            elif field_name == "volume_m3":
                v = _to_float(cell)
                if v is not None:
                    item["volume_m3"] = v
            elif field_name in ("net_weight_kg", "gross_weight_kg"):
                v = _to_float(cell)
                if v is not None:
                    item[field_name] = _maybe_convert_weight(v, header_text)
            elif field_name == "imo_flag":
                item["imo_flag"] = _is_imo_truthy(cell)

        # Skip rows that are missing the bare minimum (no ID + no dims)
        has_dims = item["length_m"] > 0 and item["width_m"] > 0 and item["height_m"] > 0
        if not item["item_id"] and not has_dims:
            skipped.append({"row": r, "reason": "no id and no dims", "raw": [str(c) for c in raw_row[:5]]})
            continue
        if not has_dims:
            skipped.append({"row": r, "reason": "missing dimensions", "raw": [str(c) for c in raw_row[:5]]})
            continue

        # Auto-fill missing item_id with a positional placeholder so the row
        # can still be packed; downstream code can flag this for the user.
        if not item["item_id"]:
            item["item_id"] = f"ITEM-{item['position'] + 1:03d}"

        # Compute volume if not provided
        if item["volume_m3"] <= 0:
            item["volume_m3"] = item["length_m"] * item["width_m"] * item["height_m"]

        # Coerce gross to net if gross missing but net present
        if not item["gross_weight_kg"] and item["net_weight_kg"]:
            item["gross_weight_kg"] = item["net_weight_kg"]

        item["raw_row_json"] = json.dumps(raw_row_dump, default=str)
        items.append(item)

    return ParseResult(
        items=items,
        header=[_normalize_header_cell(c) for c in header_row],
        column_mapping=mapping,
        skipped=skipped,
        file_type=file_type,
        sheet_name=sheet_name,
        header_row_idx=header_idx,
    )
